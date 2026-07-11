#!/usr/bin/env python3
"""
Payroll export utility — generates Excel + PDF salary sheets
grouped by company, split into FW and SPASS/SC/PR tables.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io


# ─────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────

def _thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_bottom():
    return Border(bottom=Side(style='medium'))

def _header_fill():
    return PatternFill('solid', start_color='1F3864', end_color='1F3864')

def _subheader_fill():
    return PatternFill('solid', start_color='2F5496', end_color='2F5496')

def _total_fill():
    return PatternFill('solid', start_color='D9E1F2', end_color='D9E1F2')

def _apply_border(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = _thin_border()

def _money(val):
    return round(float(val or 0), 2)


# ─────────────────────────────────────────
# MAIN EXPORT FUNCTION
# ─────────────────────────────────────────

def generate_payroll_excel(companies_data, month_str):
    """
    companies_data: list of dicts:
        {
            'name': 'TMT Quick Service Pte Ltd',
            'code': 'TQS',
            'fw_employees':   [ {company_id, name, bank_acc, net_pay}, ... ],
            'wp_employees':   [ {company_id, name, bank_acc, net_pay}, ... ]
        }
    month_str: '2026-02'  →  displayed as 'FEB -2026'
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    dt = datetime.strptime(month_str + '-01', '%Y-%m-%d')
    month_display = dt.strftime('%b').upper() + ' -' + dt.strftime('%Y')
    date_display  = 'Date ' + datetime.now().strftime('%d.%m.%Y')

    for co in companies_data:
        if not co['fw_employees'] and not co['wp_employees']:
            continue

        ws = wb.create_sheet(title=co['code'][:31])
        ws.column_dimensions['A'].width = 7   # SL NO
        ws.column_dimensions['B'].width = 12  # EMP ID
        ws.column_dimensions['C'].width = 35  # NAME
        ws.column_dimensions['D'].width = 20  # A/C
        ws.column_dimensions['E'].width = 14  # AMOUNT

        current_row = 1

        def write_company_table(employees, group_label):
            nonlocal current_row

            # Company name header
            ws.merge_cells(f'A{current_row}:E{current_row}')
            cell = ws.cell(row=current_row, column=1, value=co['name'])
            cell.font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
            cell.fill      = _header_fill()
            cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # Month subheader
            ws.merge_cells(f'A{current_row}:E{current_row}')
            cell = ws.cell(row=current_row, column=1,
                           value=f'Salary for the Month of {month_display}  [{group_label}]')
            cell.font      = Font(name='Arial', bold=True, size=11, color='FFFFFF')
            cell.fill      = _subheader_fill()
            cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # Date right-aligned
            ws.merge_cells(f'A{current_row}:E{current_row}')
            cell = ws.cell(row=current_row, column=1, value=date_display)
            cell.font      = Font(name='Arial', italic=True, size=9)
            cell.alignment = Alignment(horizontal='right')
            current_row += 1

            # Column headers
            headers = ['SL NO', 'EMP. ID', 'NAME', 'A/C', 'AMOUNT']
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=ci, value=h)
                cell.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
                cell.fill      = _subheader_fill()
                cell.alignment = Alignment(horizontal='center')
                cell.border    = _thin_border()
            current_row += 1

            # Data rows
            data_start = current_row
            for sl, emp in enumerate(employees, 1):
                ws.cell(row=current_row, column=1, value=sl).alignment      = Alignment(horizontal='center')
                ws.cell(row=current_row, column=2, value=emp.get('company_id', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=current_row, column=3, value=emp.get('name', ''))
                ws.cell(row=current_row, column=4, value=emp.get('bank_acc', ''))
                amt_cell = ws.cell(row=current_row, column=5, value=_money(emp.get('net_pay', 0)))
                amt_cell.number_format = '#,##0.00'
                for ci in range(1, 6):
                    ws.cell(row=current_row, column=ci).font   = Font(name='Arial', size=10)
                    ws.cell(row=current_row, column=ci).border = _thin_border()
                current_row += 1

            data_end = current_row - 1

            # Total row
            ws.cell(row=current_row, column=4, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
            ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='right')
            total_cell = ws.cell(row=current_row, column=5,
                                 value=f'=SUM(E{data_start}:E{data_end})')
            total_cell.font         = Font(name='Arial', bold=True, size=10)
            total_cell.number_format = '#,##0.00'
            total_cell.fill         = _total_fill()
            total_cell.border       = Border(
                top=Side(style='medium'), bottom=Side(style='double'),
                left=Side(style='thin'), right=Side(style='thin')
            )
            ws.cell(row=current_row, column=4).fill   = _total_fill()
            ws.cell(row=current_row, column=4).border = Border(
                top=Side(style='medium'), bottom=Side(style='double'),
                left=Side(style='thin'), right=Side(style='thin')
            )
            current_row += 3  # gap between tables

        if co['fw_employees']:
            write_company_table(co['fw_employees'], 'Foreign Workers')
        if co['wp_employees']:
            write_company_table(co['wp_employees'], 'SPASS / PR / Citizen')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_payroll_pdf(companies_data, month_str):
    """
    Generates a PDF using reportlab with the same layout.
    Returns BytesIO.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
    except ImportError:
        raise ImportError("reportlab not installed. Run: pip install reportlab")

    dt            = datetime.strptime(month_str + '-01', '%Y-%m-%d')
    month_display = dt.strftime('%b').upper() + ' -' + dt.strftime('%Y')
    date_display  = 'Date ' + datetime.now().strftime('%d.%m.%Y')

    dark_blue  = colors.HexColor('#1F3864')
    mid_blue   = colors.HexColor('#2F5496')
    light_blue = colors.HexColor('#D9E1F2')
    row_alt    = colors.HexColor('#EBF0FA')

    story  = []
    output = io.BytesIO()
    doc    = SimpleDocTemplate(output, pagesize=A4,
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=1.5*cm, bottomMargin=1.5*cm)

    col_widths = [1.2*cm, 2.5*cm, 7*cm, 4*cm, 3*cm]
    full_width = sum(col_widths)

    for co in companies_data:
        if not co['fw_employees'] and not co['wp_employees']:
            continue

        def build_table(employees, group_label, co_name=co['name']):
            # ── Header block (company name + month + date) ──
            hdr = [
                [co_name],
                [f"Salary for the Month of {month_display}  [{group_label}]"],
                [date_display],
            ]
            ht = Table(hdr, colWidths=[full_width])
            ht.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0), (0, 0), dark_blue),
                ('BACKGROUND',    (0, 1), (0, 1), mid_blue),
                ('TEXTCOLOR',     (0, 0), (0, 1), colors.white),
                ('FONTNAME',      (0, 0), (0, 0), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0), (0, 0), 12),
                ('FONTNAME',      (0, 1), (0, 1), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, 1), (0, 1), 10),
                ('ALIGN',         (0, 0), (0, 1), 'CENTER'),
                ('FONTNAME',      (0, 2), (0, 2), 'Helvetica-Oblique'),
                ('FONTSIZE',      (0, 2), (0, 2), 9),
                ('ALIGN',         (0, 2), (0, 2), 'RIGHT'),
                ('TOPPADDING',    (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LEFTPADDING',   (0, 0), (-1, -1), 6),
                ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
            ]))
            story.append(ht)
            story.append(Spacer(1, 4))

            # ── Data table ──
            data = [['SL NO', 'EMP. ID', 'NAME', 'A/C', 'AMOUNT']]
            total = 0
            for sl, emp in enumerate(employees, 1):
                net = _money(emp.get('net_pay', 0))
                total += net
                data.append([
                    str(sl),
                    emp.get('company_id', ''),
                    emp.get('name', ''),
                    emp.get('bank_acc', ''),
                    f"{net:,.2f}"
                ])
            data.append(['', '', '', 'TOTAL', f"{total:,.2f}"])

            row_styles = []
            for i in range(1, len(data) - 1):
                bg = colors.white if i % 2 == 1 else row_alt
                row_styles.append(('BACKGROUND', (0, i), (-1, i), bg))

            t = Table(data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                # Column header row
                ('BACKGROUND',    (0, 0), (-1, 0), mid_blue),
                ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
                ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0), (-1, 0), 9),
                ('ALIGN',         (0, 0), (-1, 0), 'CENTER'),
                # Data
                ('FONTNAME',      (0, 1), (-1, -2), 'Helvetica'),
                ('FONTSIZE',      (0, 1), (-1, -1), 9),
                # Total row
                ('BACKGROUND',    (0, -1), (-1, -1), light_blue),
                ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('LINEABOVE',     (0, -1), (-1, -1), 1.5, colors.black),
                ('LINEBELOW',     (0, -1), (-1, -1), 2,   colors.black),
                # Grid
                ('GRID',          (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN',         (4, 1), (4, -1), 'RIGHT'),
                ('ALIGN',         (0, 1), (0, -1), 'CENTER'),
                ('TOPPADDING',    (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ] + row_styles))
            story.append(t)
            story.append(Spacer(1, 20))

        if co['fw_employees']:
            build_table(co['fw_employees'], 'Foreign Workers')
        if co['wp_employees']:
            build_table(co['wp_employees'], 'SPASS / PR / Citizen')

    doc.build(story)
    output.seek(0)
    return output
